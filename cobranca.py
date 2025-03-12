import streamlit as st

import pandas as pd
import plotly.express as px
import pyodbc
import toml
from datetime import datetime, timedelta
import uuid
from st_aggrid import AgGrid, GridOptionsBuilder, JsCode, ColumnsAutoSizeMode
from PIL import Image  # Para manipulação de imagens
import df2img  # Biblioteca para converter DataFrame em imagem
import io
import traceback2 as traceback
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font

# Deve ser o primeiro comando no script
#st.set_page_config(
  #  page_title="Painel de Contas a Pagar",
   # page_icon="💳",
    #layout="wide",
    #initial_sidebar_state="expanded"
#)

# Configuração inicial do painel
def page_cobranca():
   # Paleta de cores atualizada
    COLORS = ['#13428d', '#7C3AED', '#3B82F6', '#10B981', '#EF4444', '#F59E0B']
    COLORS_DARK = ['#1b4f72', '#d35400', '#145a32', '#7b241c', '#5b2c6f']

    # Configuração do tema no session_state
    ms = st.session_state
    if "themes" not in ms:
        ms.themes = {
            "current_theme": "light",
            "light": {
                "theme.base": "light",
                "theme.backgroundColor": "#FFFFFF",  # Cor de fundo
                "theme.primaryColor": "#0095fb",     # Cor primária (botões, links)
                "theme.secondaryBackgroundColor": "#F3F4F6",  # Cor de fundo secundária
                "theme.textColor": "#0095fb",        # Cor do texto
                "button_face": "Modo Escuro 🌙",     # Texto do botão
                "colors": COLORS,                    # Paleta de cores
            },
            "dark": {
                "theme.base": "dark",
                "theme.backgroundColor": "#1F2937",  # Cor de fundo
                "theme.primaryColor": "#0095fb",     # Cor primária (botões, links)
                "theme.secondaryBackgroundColor": "#4B5563",  # Cor de fundo secundária
                "theme.textColor": "#efefef",        # Cor do texto
                "button_face": "Modo Claro 🌞",      # Texto do botão
                "colors": COLORS_DARK,               # Paleta de cores
            }
        }

    # Função para alternar o tema
    def change_theme():
        current_theme = ms.themes["current_theme"]
        ms.themes["current_theme"] = "dark" if current_theme == "light" else "light"
        ms.themes["refreshed"] = True  # Atualiza o estado

    # Configuração do tema atual
    current_theme = ms.themes["current_theme"]
    theme_config = ms.themes[current_theme]

 # Botão de alternar tema
    if st.button(theme_config["button_face"], on_click=change_theme):
        pass


    # Aplicar as cores do tema atual
    colors = theme_config["colors"]

    # Injetar CSS personalizado com base no tema atual
    st.markdown(
        f"""
        <style>
        /* ===== [CONFIGURAÇÃO GLOBAL] ===== */
        html, body, .stApp {{
            background-color: {theme_config["theme.backgroundColor"]};
            color: {theme_config["theme.textColor"]};
        }}

        /* ===== [COMPONENTES DO STREAMLIT] ===== */
        /* Ajuste para Selectbox */
        .stSelectbox > div > div {{
            background-color: {theme_config["theme.secondaryBackgroundColor"]} !important;
            color: {theme_config["theme.textColor"]} !important;
            border-radius:5px;
            border: 2px solid {theme_config["theme.primaryColor"]} !important; /* Cor oposta do tema */
        }}

        .stSelectbox > div > div:hover {{
            background-color: {theme_config["theme.primaryColor"]} !important;
            color: #FFFFFF !important;
              border: 2px solid {theme_config["theme.textColor"]} !important; /* Cor oposta do tema */
        border-radius: 5px; /* Bordas arredondadas */
        transition: border-color 0.3s ease-in-out; /* Suaviza a transição */}}
        .stMultiselect > div > div {{
            background-color: {theme_config["theme.secondaryBackgroundColor"]} !important;
            color: {theme_config["theme.textColor"]} !important;
            border-radius:5px;
            border: 2px solid {theme_config["theme.primaryColor"]} !important; /* Cor oposta do tema */
        }}

        .stMultiselec > div > div:hover {{
            background-color: {theme_config["theme.primaryColor"]} !important;
            color: #FFFFFF !important;
              border: 2px solid {theme_config["theme.textColor"]} !important; /* Cor oposta do tema */
        border-radius: 5px; /* Bordas arredondadas */
        transition: border-color 0.3s ease-in-out; /* Suaviza a transição */    
        }}

        /* Placeholder ajustado */
        .stMultiselect > div > div::placeholder {{
            color: {theme_config["theme.textColor"]} !important;
            opacity: 0.7;
        }}

        
        /* ===== [CABEÇALHOS E TÍTULOS] ===== */
        h1, h2, h3, h4, h5, h6,
        .stMarkdown h1, .stMarkdown h2, .stMarkdown h3,
        .stMarkdown h4, .stMarkdown h5, .stMarkdown h6 {{
            color: {theme_config["theme.textColor"]} !important;
        }}

    /* ===== [COMPONENTES PRINCIPAIS] ===== */
        .stDataFrame, .stMetric, .stJson, .stAlert,
        .stExpander .stMarkdown, .stTooltip, .stMetricValue, .stHeading, .stMarkdownContainer {{
            color: {theme_config["theme.textColor"]} !important;
        }}
        
        
        
        /* ===== [SIDEBAR] ===== */
        .stSidebar {{
            background-color: {theme_config["theme.secondaryBackgroundColor"]} !important;
            border-radius: 15px;
            padding: 10px;
        }}

     .nav-link.active {{
        background-color: {theme_config["theme.primaryColor"]} !important;
        color: #FFFFFF !important; /* Texto branco para contraste */
        font-weight: bold !important; /* Texto em negrito para destaque */
        border-radius: 8px; /* Bordas arredondadas */
        padding: 10px; /* Espaçamento interno */
        box-shadow: 0px 4px 6px rgba(0, 0, 0, 0.1); /* Sombra para destaque */
    }}

    /* Ícones dentro do item ativo */
    .nav-link.active .icon {{
        color: #FFFFFF !important; /* Altera a cor do ícone no item ativo */
    }}

    /* Estilo para itens inativos */
    .nav-link {{
        color: {theme_config["theme.textColor"]} !important; /* Cor do texto do tema */
        transition: background-color 0.3s, color 0.3s; /* Transição suave ao passar o mouse */
    }}

    .nav-link:hover {{
        background-color: {theme_config["theme.primaryColor"]}33; /* Cor primária translúcida */
        color: {theme_config["theme.primaryColor"]} !important; /* Texto na cor primária */
    }}
        /* Ajusta texto e fundo nos botões */
        .stButton>button {{
            background-color: {theme_config["theme.primaryColor"]} !important;
            color: #FFFFFF !important;
        }}
 /* Ajusta texto e fundo nos botões */
        .stDownloadButton>button {{
            background-color: {theme_config["theme.primaryColor"]} !important;
            color: #FFFFFF !important;
        }}
        /* ===== [GENÉRICO] ===== */
        /* Ajusta elementos dinâmicos */
        .st-emotion-cache-1cj4yv0,
        .st-emotion-cache-efbu8t {{
            background-color: {theme_config["theme.secondaryBackgroundColor"]} !important;
        
        }}
        
        /* ===== [COMPONENTES ESPECÍFICOS] ===== */
        .stMultiSelect span[data-baseweb="tag"] {{
            background-color: {theme_config["theme.primaryColor"]} !important;
            color: white !important;
        }}

        .stButton>button p {{
            color: white !important;
        }}
        
        .stDownloadButton>button p {{
            color: white !important;
        }}
        div[data-testid="stMetricValue"] {{
            color: {theme_config["theme.textColor"]} !important;
        }}

        [class*="stMetric"]
        {{
            color: {theme_config["theme.textColor"]} !important;
        }}
        
        [class*="st-emotion-cache"] {{
            color: {theme_config["theme.primaryColor"]} !important;
        }}
    .stExpander p {{
        font-size: 1.5rem !important;
            font-weight: 600;
    }}
    
  
    
    
    .stMetric {{
    border: 1px solid #e1e4e8;
    border-radius: 8px;
    padding: 15px;
    background-color: #f8f9fa;
    margin: 10px 0;
    }}
    .stMetric label {{
        font-size: 1.1rem !important;
        font-weight: 500;
        color: #666;
    }}
    .stMetric div {{
        font-size: 1.4rem !important;
        font-weight: 600;
        color: #2c3e50;
    }}
        """,
        unsafe_allow_html=True
    )

    def load_db_config():
        config = toml.load("config.toml")
        return config["database"]

    @st.cache_data(ttl=300)
    def fetch_data():
        db_config = load_db_config()
        cnxn = pyodbc.connect(f'DRIVER={{ODBC Driver 17 for SQL Server}};'
                              f'SERVER={db_config["server"]};DATABASE={db_config["database"]};'
                              f'UID={db_config["username"]};PWD={db_config["password"]}')
        query = """SELECT 
    cf.NomeCliFor AS [Cliente],
    cf.RazaoSocial AS [Razão Social],
    ccf.Denominacao AS [Categoria CliFor],
    cf.CPFCGCCLIFOR AS [CFP/CNPJ],
    CASE WHEN cf.TipoPessoa = 'F' THEN 'Fisica' ELSE 'Juridica' END AS [Pessoa],
    p.NomePortador AS [Portador],
    c.NUMDOC AS [NUMDOC],
    c.HISTORICO AS [Historico],
    CAST(c.DTEMISSAO AS date) AS [Dt Emissão],
    CAST(c.DTVENCIMENTO AS date) AS [Dt Vencimento],
    CAST(c.DTBAIXA AS date) AS [Dt Baixa],
    CASE 
        WHEN c.DTBAIXA IS NULL AND CAST(c.DTVENCIMENTO AS date) = CAST(GETDATE() AS date) THEN 'Não Pago - Vence Hoje'
        WHEN c.DTBAIXA IS NOT NULL AND CAST(c.DTBAIXA AS date) <= CAST(c.DTVENCIMENTO AS date) THEN 'Pago no Prazo'
        WHEN c.DTBAIXA IS NOT NULL AND CAST(c.DTBAIXA AS date) > CAST(c.DTVENCIMENTO AS date) THEN 'Pago em Atraso'
        WHEN c.DTBAIXA IS NULL AND CAST(c.DTVENCIMENTO AS date) < CAST(GETDATE() AS date) THEN 'Não Pago - Em Atraso'
        WHEN c.DTBAIXA IS NULL AND CAST(c.DTVENCIMENTO AS date) > CAST(GETDATE() AS date) THEN 'Não Pago'
  
        ELSE 'Indefinido'
    END AS [Status Pagamento],
    CASE 
        WHEN c.DTBAIXA IS NULL AND CAST(c.DTVENCIMENTO AS date) < CAST(GETDATE() AS date) 
            THEN DATEDIFF(DAY, CAST(c.DTVENCIMENTO AS date), CAST(GETDATE() AS date))
        WHEN c.DTBAIXA IS NOT NULL AND CAST(c.DTBAIXA AS date) > CAST(c.DTVENCIMENTO AS date) 
            THEN DATEDIFF(DAY, CAST(c.DTVENCIMENTO AS date), CAST(c.DTBAIXA AS date))
        ELSE 0 
    END AS [Dias Atraso],
    c.ValorTitulo AS [VR Nominal],
    dbCronos.dbo.fn_valorcpr(c.IdCPR, 'C', c.DTVENCIMENTO, GETDATE()) AS [VR Corrigido],
    td.TipoDoc AS [Tipo Documento],
    f.NomeFilial AS [Empresa - Filial],
    pc.Denominacao AS [Plano de Contas],
    CASE pc.TipoConta 
        WHEN 'C' THEN 'Custo'
        WHEN 'V' THEN 'Despesa Variável'
        WHEN 'D' THEN 'Despesa Fixa'
        WHEN 'R' THEN 'Receita Operacional'
        WHEN 'O' THEN 'Compensação'
        ELSE '' 
    END AS [Plano de Contas TP],
    dbCronos.dbo.fn_DscMeioPagCPR(c.IdCPR) AS [Meio de Pagamento]
FROM dbCronos.dbo.CPR c
INNER JOIN dbCronos.dbo.Cli_For cf ON c.CodCliFor = cf.CodCliFor
INNER JOIN dbCronos.dbo.CategoriaCliFor ccf ON cf.IdCategoriaCliFor = ccf.IdCategoriaCliFor
INNER JOIN dbCronos.dbo.Portador p ON c.CodPortador = p.CodPortador
INNER JOIN dbCronos.dbo.TipoDoc td ON c.CodTipoDoc = td.CodTipoDoc
INNER JOIN dbCronos.dbo.PlanoConta pc ON c.IdPlanoConta = pc.IdPlanoConta
INNER JOIN dbCronos.dbo.Filiais f ON c.CodFilial = f.CodFilial
WHERE c.PagRec = 'R'
    """
        data = pd.read_sql(query, con=cnxn)
        return data



    def safe_to_datetime(series):
        return pd.to_datetime(series, errors='coerce')

    data = fetch_data()
    if data.empty:
        st.error("Nenhum dado foi carregado. Verifique sua consulta SQL.")
        return

    data['Dt Emissão'] = safe_to_datetime(data['Dt Emissão'])
    data['Dt Vencimento'] = safe_to_datetime(data['Dt Vencimento'])
    data['Ano'] = data['Dt Vencimento'].dt.year.dropna().astype(int)

    month_translation = {
        'January': 'janeiro', 'February': 'fevereiro', 'March': 'março',
        'April': 'abril', 'May': 'maio', 'June': 'junho',
        'July': 'julho', 'August': 'agosto', 'September': 'setembro',
        'October': 'outubro', 'November': 'novembro', 'December': 'dezembro'
    }
    month_order = list(month_translation.values())
    data['Mês'] = data['Dt Vencimento'].dt.strftime('%B').map(month_translation).fillna('Indefinido')
    data['Mês'] = pd.Categorical(data['Mês'], categories=month_order, ordered=True)
    data['Semana'] = data['Dt Vencimento'].dt.isocalendar().week

    # Função para calcular as datas de início e fim das semanas de um ano específico
    def calcular_periodo_semanas_por_ano(ano):
        semanas = []
        # Encontrar a primeira segunda-feira do ano
        primeira_data = datetime(ano, 1, 1)
        primeira_segunda = primeira_data + timedelta(days=(7 - primeira_data.weekday()) % 7)

        for semana in range(1, 54):  # Até 53 semanas possíveis
            inicio_semana = primeira_segunda + timedelta(weeks=semana - 1)
            fim_semana = inicio_semana + timedelta(days=6)
            # Parar se já ultrapassamos o ano
            if inicio_semana.year > ano:
                break
            semanas.append({
                "Semana": semana,
                "Inicio": inicio_semana,
                "Fim": fim_semana,
                "Periodo": f"{inicio_semana.strftime('%d/%m')} - {fim_semana.strftime('%d/%m')}"
            })
        return semanas
    
# Mapeamento completo da coluna "Plano de Contas" para as categorias de despesa
    mapeamento = {
        'Impostos': [
            'ICMS Normal', 'ICMS Substituto', 'ICMS Antecipado', 'PIS - 8109', 'Cofins - 2172',
            'IRPJ', 'CSLL', 'ISS', 'IPTU', 'IOF', 'Contribuição Social', 'Contribuição Sindical',
            'Taxa Incêndios', 'Taxa Frete Comodato', 'Taxa de Entrega (Frete)', 'Multas de ICMS',
            'Parcelamento Dívida Ativa - 1124', 'IRRF - INSS do Prolabore', 'IRRF - INNS Funcionários',
            'Parcelamento Simples Nacional', 'Parcelamento ICMS', 'Parcelamento PIS', 'Parcelamento COFINS',
            'Parcelametos Previdenciário Demais Débitos', 'TCFA - TAXA AMBIENTAL', 'IPI - Imposto sobre Produtos Industrializados',
            'Imposto Retido', 'DAS / Simples Nacional', 'Taxas de Certificações e Registros', 'GNRE',
            'GRU Judicial', 'DAM - não usar mais', 'DAE - não usar mais', 'DARF * não usar mais esse', 
        'TLLF', 'Parcelamento ISS', 'CPMF'
        ],
        'RH': [
            'Salários', 'Gratificações', 'Distribuição de Lucros','Férias', '13º Salário', 'PRO LABORE', 'Vale Transporte',
            'Inss / Gps', 'Fgts', 'Rescisões e Indenizações', 'Adiantamento de Salário', 'Adianatamento Salarial Fora da Folha',
            'Hora Extra', 'Vale Refeições', 'Assistencia Médica', 'Assistencia Odontologica', 'Farmácia',
            'Lazer e Confraternizações', 'Formação Profissional', 'Estágio', 'Treinamento', 'Fardamento',
            'Exames admissionais e demissionais', 'Seguro de Vida', 'Material de EPI', 'Ajuda de Custo',
            'CIPA', 'Aniversariante do Mês', 'coffee break', 'Endomarketing', 'Cursos e treinamentos', 'Diretor', 'Sócios', 'Emprestimo por fora da folha'
        ],
        'Pagamento a Fornecedores': [
            'Fornecedores Mercantil', 'Fretes Compra', 'Frete Vendas Mercantil', 'Frete Grátis Comodato',  
            'Frete Pago Cliente', 'Frete Serviço', 'Compra de Produto', 'Devolução de Compra', 'Adiantamento a Fornecedor',  
            'Prestador de Serviço PF', 'Prestador de Serviço PJ', 'Comissão Serviço', 'Comissão Vendas', 'Comissao PJ',  
            'Equipamentos de Informatica', 'Máquinas e Equipamentos', 'Aquisição de Software', 'Móveis e Equipamentos',  
            'Material de Informática', 'Uso e Consumo Diversos', 'Embalagens', 'Equi. transporte de Material',  
            'Aluguel de Equipamentos', 'Consórcios Bancários', 'Consórcio de Imovél', 'Taxa Adm de Consorcio',  
            'Nota de Credito', 'Bonificação', 'Prémio de vendas', 'Reembolso', 'Empréstimo por antecipados',  
            'Operação de Crédito BNDS', 'Frete Grátis E-commerce'  

        ],
       'Pagamentos Administrativos': [
            'Telefone', 'Cia de Aguas e Esgotos', 'Energia Elétrica', 'Aluguel do Imóvel', 'Material Expediente, Papelaria',
            'Matérial de Limpeza', 'Honorários Contabeis', 'Honorários Jurídicos', 'Honorários Sistema', 'Honorários Ti',
            'Assinaturas de Jornais / Revistas', 'Internet', 'Manutenção Predial', 'Seguro de Imóveis', 'Serviço de TV',
            'Serviços de Segurança', 'Correios / Sedex', 'Despesas Viagens Vendedor Externo', 'Despesas Plotagens / Xerox',
            'Publicidade', 'Patrocinio', 'Eventos Realizados', 'Brindes', 'Divulgação Rede Social', 'Carro de Som',
            'Panfletos', 'Eventos', 'Manutenção Eletrodomésticos e Eletroeletrônicos', 'Reformas e Construções',
            'Reparos e Consertos de Bens e Moveis', 'Predios', 'Construções e Reformas', 'Combustivel', 'Emplacamento',
            'Ipva', 'Multas de Trânsito', 'Peça e Serviço para Veiculo', 'Seguro de Veiculo', 'Leasing de Veiculo',
            'Consórcio de Veiculo', 'Lavagem', 'Estacionamento', 'Pedágio', 'Viagens Vendedor', 'Viagens Geral Colaborador',
            'Combustível Vendedores', 'Alimentação Vendedores', 'Hospedagem Vendedor', 'Hospedagem Colaborador',
            'Combustível Colaborador', 'Alimentação Colaborador', 'Passagens', 'Uber/Aplicativos', 'Deslocamento Corporativo / Outros',
            'Mat.Higiene Limpeza e Mat. Copa e Cozinha', 'Autenticação de Firma', 'Alvará', 'Taxa Boleto Não Pago',
            'Tarifa Bancaria', 'Despesa de Cobrança Bancaria', 'Taxa Administrativa de Cartão', 'Serasa / SPC', 'Valores SKY',
            'NF´S SKY', 'Gastos RDVC SKY', 'Reembolso RDVC SKY', 'Crédito Voucher', 'Recarga', 'Fatura Cartão de Crédito',
            'CARTÃO COPORATIVO', 'Animador', 'Outras Compensações'
        ]
    }
    # Função para categorizar o "Plano de Contas"
    def categorizar(plano_contas):
        for categoria, itens in mapeamento.items():
            if plano_contas in itens:
                return categoria
        return 'Outros'  # Para itens não mapeados

    # Aplicar o mapeamento (assumindo que 'data' e 'mapeamento' já estão definidos anteriormente no código)
    data['Categoria Despesa'] = data['Plano de Contas'].apply(categorizar)

    # Sidebar para filtros
    st.sidebar.title("Filtros")
    ano_atual = datetime.now().year
    first_day_of_year = datetime(datetime.today().year, 1, 1).date()

    # Filtro de Dia (ou intervalo de datas)
    dias_selecionados = st.sidebar.date_input(
        "Filtrar por Dia(s)",
        value=[first_day_of_year, datetime.today().date()],
        format="DD/MM/YYYY",
        min_value=data['Dt Vencimento'].min().date(),
        max_value=data['Dt Vencimento'].max().date(),
        label_visibility="collapsed"
    )

    # Aplicar o filtro de dia
    filtered_data = data.copy()  # Criar uma cópia para evitar modificar o DataFrame original
    if len(dias_selecionados) == 1:
        filtered_data = filtered_data[pd.to_datetime(filtered_data['Dt Vencimento']).dt.date == dias_selecionados[0]]
    else:
        filtered_data = filtered_data[(pd.to_datetime(filtered_data['Dt Vencimento']).dt.date >= dias_selecionados[0]) &
                                    (pd.to_datetime(filtered_data['Dt Vencimento']).dt.date <= dias_selecionados[1])]

    # Filtro de Ano usando os anos disponíveis nos dados
    anos_disponiveis = ['Todos'] + sorted(filtered_data['Ano'].dropna().astype(int).unique().tolist())
    ano_selecionado = st.sidebar.selectbox(
        "Selecione o Ano",
        anos_disponiveis,
        index=anos_disponiveis.index(ano_atual) if ano_atual in anos_disponiveis else 0
    )

    # Filtragem dos dados por ano
    if ano_selecionado != 'Todos':
        filtered_data = filtered_data[filtered_data['Ano'] == ano_selecionado]

    # Calcula as semanas do ano selecionado
    if ano_selecionado != 'Todos':
        semanas_do_ano = calcular_periodo_semanas_por_ano(ano_selecionado)
        semanas_opcoes = ["Todos"] + [f"Semana {s['Semana']} ({s['Periodo']})" for s in semanas_do_ano]
    else:
        semanas_opcoes = ["Todos"]

    # Filtro de Semana
    semana_selecionada = st.sidebar.selectbox("Selecione a Semana", semanas_opcoes)

    # Filtro de Status de Pagamento com multiselect, incluindo "Todos" como padrão
    status_pagamento = st.sidebar.multiselect(
        "Filtrar por Status de Pagamento",
        ['Todos', 'Não Pago', 'Não Pago - Vence Hoje','Não Pago - Em Atraso', 'Pago em Atraso', 'Pago no Prazo', 'Indefinido'],
        default=['Todos'],
        key="status_pagamento"
    )

    # Filtro de Categoria Despesa com multiselect, incluindo "Todos" como padrão
    despesas_disponiveis = ['Todos'] + sorted(filtered_data['Categoria Despesa'].dropna().unique().tolist())
    despesa_selecionada = st.sidebar.multiselect(
        "Filtrar por Tipo de Despesas",
        despesas_disponiveis,
        default=['Todos'],
        key="despesa_selecionada"
    )

    # Filtros adicionais
    mes_selecionado = st.sidebar.selectbox("Filtrar por Mês", ['Todos'] + list(filtered_data['Mês'].cat.categories))
    filtro_x = st.sidebar.selectbox("Filtrar eixo X", ['Plano de Contas', 'Plano de Contas TP', 'Portador', 'Tipo Documento', 'Meio de Pagamento', 'Categoria Despesa'])
    # Filtro simplificado para gráfico principal
    categoria_grafico = st.sidebar.selectbox(
        "Selecione o Gráfico Principal",
        ['Mês', 'Plano de Contas TP', 'Portador', 'Tipo Documento', 'Categoria Despesa']
    )

    # Filtragem dos dados
    if semana_selecionada != 'Todos':
        semana_numero = int(semana_selecionada.split()[1])
        filtered_data = filtered_data[filtered_data['Semana'] == semana_numero]
    if mes_selecionado != 'Todos':
        filtered_data = filtered_data[filtered_data['Mês'] == mes_selecionado]
    if status_pagamento != ['Todos']:
        filtered_data = filtered_data[filtered_data['Status Pagamento'].isin(status_pagamento)]
    if despesa_selecionada != ['Todos']:
        filtered_data = filtered_data[filtered_data['Categoria Despesa'].isin(despesa_selecionada)]

    # Sumário mensal e semanal
    monthly_summary = filtered_data.groupby([filtro_x, 'Mês']).agg({'VR Nominal': 'sum'}).reset_index()
    weekly_summary = filtered_data.groupby([filtro_x, 'Semana']).agg({'VR Nominal': 'sum'}).reset_index()

    # Obtenção das datas de hoje, começo da semana e do mês
    today = datetime.today().date()
    start_of_week = today - timedelta(days=today.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    start_of_month = today.replace(day=1)
    end_of_month = (start_of_month + timedelta(days=31)).replace(day=1) - timedelta(days=1)

    # Filtragem por hoje, semana e mês
    today_data = filtered_data[pd.to_datetime(filtered_data['Dt Vencimento']).dt.date == today]
    week_data = filtered_data[(pd.to_datetime(filtered_data['Dt Vencimento']).dt.date >= start_of_week) &
                            (pd.to_datetime(filtered_data['Dt Vencimento']).dt.date <= end_of_week)]
    month_data = filtered_data[(pd.to_datetime(filtered_data['Dt Vencimento']).dt.date >= start_of_month) &
                            (pd.to_datetime(filtered_data['Dt Vencimento']).dt.date <= end_of_month)]

    # Total mensal para o gráfico
    total_monthly = monthly_summary.groupby('Mês')['VR Nominal'].transform('sum')
    total_weekly = weekly_summary.groupby('Semana')['VR Nominal'].transform('sum')
    monthly_summary['Percentual'] = (monthly_summary['VR Nominal'] / total_monthly) * 100
    weekly_summary['Percentual'] = (weekly_summary['VR Nominal'] / total_weekly) * 100

    # Agrupa os dados para o resumo mensal
    resumo_mensal = filtered_data.groupby(["Mês"]).agg({"VR Nominal": "sum"}).reset_index()
    resumo_mensal["Mês"] = pd.Categorical(resumo_mensal["Mês"], categories=month_order, ordered=True)
    resumo_mensal = resumo_mensal.sort_values("Mês")

    if "VR Nominal" in resumo_mensal.columns:
        total_valor = resumo_mensal["VR Nominal"].sum()
        resumo_mensal["Percentual"] = (resumo_mensal["VR Nominal"] / total_valor) * 100
    else:
        st.error("A coluna 'VR Nominal' está ausente no resumo mensal.")
        return

    # Sumário por categoria de despesa
    categoria_summary = filtered_data.groupby('Categoria Despesa').agg({'VR Nominal': 'sum'}).reset_index()
    total_categoria = categoria_summary['VR Nominal'].sum()
    categoria_summary['Percentual'] = (categoria_summary['VR Nominal'] / total_categoria) * 100

    # Determinar o período para o título com base nos filtros aplicados
    if len(dias_selecionados) == 1:
        periodo = f"em {dias_selecionados[0].strftime('%d/%m/%Y')}"
    else:
        periodo = f"de {dias_selecionados[0].strftime('%d/%m/%Y')} até {dias_selecionados[1].strftime('%d/%m/%Y')}"

    # Exibir resumo por categoria com período no título
    st.subheader(f"Resumo por Categoria de Despesa ({periodo})")
    st.dataframe(
        categoria_summary.style.format({
            'VR Nominal': 'R${:,.2f}',
            'Percentual': '{:.1f}%'
        }),
        hide_index=True
    )

    show_graphs = st.checkbox("Exibir Gráficos", value=True)

    if show_graphs:
        try:
            # Gráfico principal
            st.subheader(f"Gráfico de {categoria_grafico}")
            if not filtered_data.empty:
                if categoria_grafico == 'Mês':
                    grouped_data = filtered_data.groupby('Mês').agg({'VR Nominal': 'sum'}).reset_index()
                    fig = px.bar(
                        grouped_data,
                        x='Mês',
                        y='VR Nominal',
                        text=grouped_data['VR Nominal'].map(lambda x: f"R${x:,.2f}"),
                        title=f'Total por Mês',
                        color='Mês',
                        barmode='group'
                    )
                else:
                    grouped_data = filtered_data.groupby(categoria_grafico).agg({'VR Nominal': 'sum'}).reset_index()
                    fig = px.bar(
                        grouped_data,
                        x=categoria_grafico,
                        y='VR Nominal',
                        text=grouped_data['VR Nominal'].map(lambda x: f"R${x:,.2f}"),
                        title=f'Total por {categoria_grafico}',
                        color=categoria_grafico,
                        barmode='group'
                    )
                st.plotly_chart(fig, key='grafico_principal', use_container_width=True)

            # Visões Adicionais em expansor
            with st.expander("Visões Adicionais"):
                tab1, tab2 = st.tabs(["Outras Categorias", "Resumo Mensal"])
                with tab1:
                    for categoria in ['Plano de Contas TP', 'Portador', 'Tipo Documento', 'Meio de Pagamento']:
                        if categoria != categoria_grafico:  # Evitar redundância com o gráfico principal
                            summary = filtered_data.groupby(categoria).agg({'VR Nominal': 'sum'}).reset_index()
                            fig = px.bar(
                                summary,
                                x=categoria,
                                y='VR Nominal',
                                text=summary['VR Nominal'].map(lambda x: f"R${x:,.2f}"),
                                title=f'Total por {categoria}',
                                color=categoria,
                                barmode='group'
                            )
                            st.plotly_chart(fig, key=f'visoes_{categoria}', use_container_width=True)
                with tab2:
                    resumo_mensal = filtered_data.groupby("Mês").agg({"VR Nominal": "sum"}).reset_index()
                    resumo_mensal["Percentual"] = (resumo_mensal["VR Nominal"] / resumo_mensal["VR Nominal"].sum()) * 100
                    fig_abs = px.bar(
                        resumo_mensal,
                        x="Mês",
                        y="VR Nominal",
                        text=resumo_mensal["VR Nominal"].map(lambda x: f"R${x:,.2f}"),
                        title="Distribuição Absoluta",
                        color="Mês",
                        barmode='group'
                    )
                    fig_perc = px.bar(
                        resumo_mensal,
                        x="Mês",
                        y="Percentual",
                        text=resumo_mensal["Percentual"].map(lambda x: f"{x:.1f}%"),
                        title="Distribuição Percentual",
                        color="Mês",
                        barmode='group'
                    )
                    st.plotly_chart(fig_abs, key='resumo_mensal_abs', use_container_width=True)
                    st.plotly_chart(fig_perc, key='resumo_mensal_perc', use_container_width=True)

        except Exception as e:
            st.error(f"Erro ao gerar gráficos: {str(e)}")

        # Função para formatação de moeda
    def format_currency(value):
        return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    # Função para formatação de data
    def format_date(date):
        return date.strftime('%d/%m/%Y') if pd.notnull(date) else ''

    # Função para aplicar cores com base no status de pagamento
    def get_row_style():
        return JsCode("""
            function(params) {
                if (params.data['Status Pagamento'] === 'Pago no Prazo') {
                    return {'backgroundColor': '#2ecc71', 'color': 'white'};
                } else if (params.data['Status Pagamento'] === 'Pago em Atraso') {
                    return {'backgroundColor': '#538405', 'color': 'white'};
                } else if (params.data['Status Pagamento'] === 'Não Pago - Em Atraso') {
                    return {'backgroundColor': '#e74c3c', 'color': 'white'};
                } else if (params.data['Status Pagamento'] === 'Não Pago') {
                    return {'backgroundColor': '#f1a11f', 'color': 'black'};
                } else if (params.data['Status Pagamento'] === 'Não Pago - Vence Hoje') {
                    return {'backgroundColor': '#f8e620', 'color': 'black'};
                } else {
                    return {'backgroundColor': '#d3d3d3', 'color': 'black'};
                }
            };
        """)

    # Função para gerar a legenda das cores
    def display_color_legend():
        legend_html = """
        <div style="margin-bottom: 10px;">
            <strong>Legenda de Cores:</strong>
            <ul style="list-style-type: none; padding: 0;">
                <li><span style="display: inline-block; width: 20px; height: 20px; background-color: #2ecc71; margin-right: 10px;"></span> Pago no Prazo</li>
                <li><span style="display: inline-block; width: 20px; height: 20px; background-color: #538405; margin-right: 10px;"></span> Pago em Atraso</li>
                <li><span style="display: inline-block; width: 20px; height: 20px; background-color: #e74c3c; margin-right: 10px;"></span> Não Pago - Em Atraso</li>
                <li><span style="display: inline-block; width: 20px; height: 20px; background-color: #f1a11f; margin-right: 10px;"></span> Não Pago</li>
                <li><span style="display: inline-block; width: 20px; height: 20px; background-color: #f8e620; margin-right: 10px;"></span> Não Pago - Vence Hoje</li>
                <li><span style="display: inline-block; width: 20px; height: 20px; background-color: #d3d3d3; margin-right: 10px;"></span> Indefinido</li>
            </ul>
        </div>
        """
        st.markdown(legend_html, unsafe_allow_html=True)

    # Função para aplicar estilos ao DataFrame
    def apply_styles(df):
        def color_rows(row):
            status = row['Status Pagamento']
            if status == 'Pago no Prazo':
                return ['background-color: #2ecc71; color: white'] * len(row)
            elif status == 'Pago em Atraso':
                return ['background-color: #538405; color: white'] * len(row)
            elif status == 'Não Pago - Em Atraso':
                return ['background-color: #e74c3c; color: white'] * len(row)
            elif status == 'Não Pago':
                return ['background-color: #f1a11f; color: black'] * len(row)
            elif status == 'Não Pago - Vence Hoje':
                return ['background-color: #f8e620; color: black'] * len(row)
            else:
                return ['background-color: #d3d3d3; color: black'] * len(row)

        styled_df = df.style.apply(color_rows, axis=1).set_properties(**{
            'font-size': '12pt',
            'text-align': 'center',
            'border-style': 'solid',
            'border-width': '1px',
            'border-color': '#000000'
        }).set_table_styles([
            {'selector': 'th',
            'props': [('font-weight', 'bold'),
                    ('background-color', '#D3D3D3'),
                    ('text-align', 'center'),
                    ('border-style', 'solid'),
                    ('border-width', '1px'),
                    ('border-color', '#000000')]}
        ])
        return styled_df

    # Função para converter DataFrame estilizado em imagem PNG
    def dataframe_to_png(df):
        try:
            if df.empty:
                st.warning("O DataFrame está vazio. Não é possível gerar uma imagem.")
                return None

            # Aplicar estilos ao DataFrame
            styled_df = apply_styles(df)

            # Converter o DataFrame estilizado para HTML
            html = styled_df.to_html()

            # Criar uma figura com Matplotlib
            num_cols = len(df.columns)
            num_rows = len(df)
            fig_width = max(10, num_cols * 2.5)  # Largura ajustada para colunas expandidas
            fig_height = max(4, num_rows * 0.8)  # Altura ajustada para linhas
            fig, ax = plt.subplots(figsize=(fig_width, fig_height))

            # Remover eixos
            ax.axis('off')

            # Renderizar a tabela como texto (usando HTML temporariamente para estilos)
            table = ax.table(cellText=df.values,
                            colLabels=df.columns,
                            cellLoc='center',
                            loc='center',
                            bbox=[0, 0, 1, 1])

            # Ajustar o tamanho da fonte e escala
            table.auto_set_font_size(False)
            table.set_fontsize(12)
            table.scale(1.5, 1.5)  # Aumentar o tamanho das células

            # Aplicar cores às células
            for i, row in enumerate(df.itertuples()):
                status = row._asdict().get('Status Pagamento', '')
                if status == 'Pago no Prazo':
                    color = '#2ecc71'
                    text_color = 'white'
                elif status == 'Pago em Atraso':
                    color = '#538405'
                    text_color = 'white'
                elif status == 'Não Pago - Em Atraso':
                    color = '#e74c3c'
                    text_color = 'white'
                elif status == 'Não Pago':
                    color = '#f1a11f'
                    text_color = 'black'
                elif status == 'Não Pago - Vence Hoje':
                    color = '#f8e620'
                    text_color = 'black'
                else:
                    color = '#d3d3d3'
                    text_color = 'black'
                for j in range(len(df.columns)):
                    table[(i + 1, j)].set_facecolor(color)
                    table[(i + 1, j)].set_text_props(color=text_color)
                    table[(i + 1, j)].set_edgecolor('#000000')  # Bordas pretas

            # Estilizar o cabeçalho
            for j in range(len(df.columns)):
                table[(0, j)].set_facecolor('#D3D3D3')
                table[(0, j)].set_text_props(weight='bold', color='black')
                table[(0, j)].set_edgecolor('#000000')

            # Ajustar largura das colunas automaticamente
            table.auto_set_column_width(col=list(range(len(df.columns))))

            # Salvar a imagem em um buffer
            buf = io.BytesIO()
            plt.savefig(buf, format='png', bbox_inches='tight', dpi=150, facecolor='white')
            buf.seek(0)
            plt.close(fig)

            st.success("Imagem PNG gerada com sucesso!")
            return buf

        except Exception as e:
            st.error(f"Erro ao converter DataFrame para PNG: {str(e)}")
            st.error(traceback.format_exc())
            return None

    # Função para exibir a tabela com AgGrid (mantida como no original)
    def display_aggrid_table(title, df, key):
        st.subheader(title)
        if show_legend:
            display_color_legend()
        if df.empty:
            st.write("Nenhuma conta encontrada.")
            return
        
        df_display = df.copy()
        if 'Dt Vencimento' in df_display.columns:
            df_display['Dt Vencimento'] = pd.to_datetime(df_display['Dt Vencimento'], errors='coerce').apply(lambda d: format_date(d) if pd.notnull(d) else '')
        if 'Dt Emissão' in df_display.columns:
            df_display['Dt Emissão'] = pd.to_datetime(df_display['Dt Emissão'], errors='coerce').apply(lambda d: format_date(d) if pd.notnull(d) else '')
        if 'Dt Baixa' in df_display.columns:
            df_display['Dt Baixa'] = pd.to_datetime(df_display['Dt Baixa'], errors='coerce').apply(lambda d: format_date(d) if pd.notnull(d) else '')
        if 'VR Nominal' in df_display.columns:
            df_display['VR Nominal'] = df_display['VR Nominal'].apply(format_currency)
        if 'VR Corrigido' in df_display.columns:
            df_display['VR Corrigido'] = df_display['VR Corrigido'].apply(format_currency)
        
        gb = GridOptionsBuilder.from_dataframe(df_display)
        gb.configure_selection('multiple', use_checkbox=True)
        gb.configure_grid_options(getRowStyle=get_row_style())
        gb.configure_column('Cliente', headerCheckboxSelection=True)
        gb.configure_pagination(enabled=True, paginationAutoPageSize=False, paginationPageSize=50)
        
        grid_options = gb.build()
        grid_options['domLayout'] = 'normal'
        grid_options['rowHeight'] = 30
        altura_tabela = 650
        
        ms = st.session_state
        if "themes" not in ms:
            st.error("Tema não configurado no session_state.")
            return
        current_theme = ms.themes["current_theme"]
        theme_config = ms.themes[current_theme]

        if current_theme == "dark":
            aggrid_theme = 'streamlit'
            grid_options['customCss'] = {
                'backgroundColor': theme_config["theme.backgroundColor"],
                'color': theme_config["theme.textColor"],
                'headerBackgroundColor': theme_config["theme.secondaryBackgroundColor"],
                'gridLineColor': '#4B5563',
                'selectionBackgroundColor': f'{theme_config["theme.primaryColor"]}33',
            }
        else:
            aggrid_theme = 'streamlit'
            grid_options['customCss'] = {
                'backgroundColor': theme_config["theme.backgroundColor"],
                'color': theme_config["theme.primaryColor"],
                'headerBackgroundColor': theme_config["theme.secondaryBackgroundColor"],
                'gridLineColor': '#E5E7EB',
                'selectionBackgroundColor': f'{theme_config["theme.primaryColor"]}33',
            }
        
        st.markdown(
            f"""
            <style>
            .ag-theme-streamlit {{
                --ag-background-color: {theme_config["theme.backgroundColor"]};
                --ag-foreground-color: {theme_config["theme.primaryColor"]};
                --ag-header-background-color: {theme_config["theme.secondaryBackgroundColor"]};
                --ag-grid-line-color: {'#4B5563' if current_theme == "dark" else '#E5E7EB'};
                --ag-row-hover-color: {theme_config["theme.primaryColor"]}33;
                --ag-selected-row-background-color: {theme_config["theme.primaryColor"]}66;
            }}
            .ag-header-cell {{
                color: {theme_config["theme.textColor"]} !important;
            }}
            </style>
            """,
            unsafe_allow_html=True
        )
        
        grid_response = AgGrid(
            df_display,
            gridOptions=grid_options,
            height=altura_tabela,
            allow_unsafe_jscode=True,
            theme=aggrid_theme,
            update_mode='SELECTION_CHANGED',
            columns_auto_size_mode=ColumnsAutoSizeMode.FIT_CONTENTS,
            key=key
        )
        
        selected_rows = grid_response.get('selected_rows', [])
        selected_df = pd.DataFrame(selected_rows)
        
        if not selected_df.empty:
            st.write("### Tabela Auxiliar - Linhas Selecionadas")
            st.dataframe(selected_df, hide_index=True)
            
            if 'VR Corrigido' in selected_df.columns:
                def convert_value(x):
                    if isinstance(x, str) and x.strip():
                        return float(x.replace('R$ ', '').replace('.', '').replace(',', '.'))
                    return 0.0
                total_corrigido = selected_df['VR Corrigido'].apply(convert_value).sum()
                st.metric(label="Total Selecionado💲🪙", value=format_currency(total_corrigido))
            
            st.write("### Download da Tabela Selecionada")
            img_buffer = dataframe_to_png(selected_df)
            if img_buffer:
                st.download_button(
                    label="Download Imagem 🖼️",
                    data=img_buffer,
                    file_name=f"{title}_selecionados.png",
                    mime="image/png",
                    key=f"{key}_download_image"  # Correção: usando f-string para chave única
                )
            else:
                st.warning("Não foi possível gerar o arquivo PNG para download.")
                
            # Gerar arquivo Excel
            excel_buffer = io.BytesIO()
            selected_df.to_excel(excel_buffer, index=False, engine='openpyxl')
            excel_buffer.seek(0)

            # Carregar o workbook com openpyxl
            wb = openpyxl.load_workbook(excel_buffer)
            ws = wb.active

            # Criar uma tabela estilizada
            tab = Table(displayName="TabelaSelecionada", ref=f"A1:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}")
            style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            tab.tableStyleInfo = style
            ws.add_table(tab)

            # Encontrar a coluna 'Status Pagamento'
            status_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(1, col).value == 'Status Pagamento':
                    status_col = col
                    break

            # Aplicar formatação condicional se a coluna for encontrada
            if status_col:
                color_map = {
                    'Pago no Prazo': '2ECC71',          # Verde
                    'Pago em Atraso': '538405',         # Verde escuro
                    'Não Pago - Em Atraso': 'E74C3C',   # Vermelho
                    'Não Pago': 'F1A11F',              # Laranja
                    'Não Pago - Vence Hoje': 'F8E620',  # Amarelo
                    'Indefinido': 'D3D3D3'             # Cinza
                }

                for status, color in color_map.items():
                    fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    font_color = 'FFFFFF' if status in ['Pago no Prazo', 'Pago em Atraso', 'Não Pago - Em Atraso'] else '000000'
                    font = Font(color=font_color)
                    rule = openpyxl.formatting.rule.CellIsRule(
                        operator='equal',
                        formula=[f'"{status}"'],
                        fill=fill,
                        font=font
                    )
                    ws.conditional_formatting.add(f"A2:{openpyxl.utils.get_column_letter(ws.max_column)}{ws.max_row}", rule)

            # Salvar o arquivo no buffer
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            # Botão de download no Streamlit
            st.download_button(
                label="Download Excel 📊",
                data=excel_buffer,
                file_name="tabela_auxiliar.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"{key}_download_excel"  # Correção: usando f-string para chave única
            )
        else:
            st.info("Nenhuma linha selecionada. Selecione linhas na tabela acima para ver a tabela auxiliar e o botão de download.")
    # Configuração de datas
   # today = pd.to_datetime('today').normalize().date()
    #start_of_week = (pd.to_datetime('today').normalize() - pd.offsets.Week(weekday=0)).date()
    #end_of_week = (start_of_week + pd.Timedelta(days=6)).date()
    #start_of_month = pd.to_datetime('today').normalize().replace(day=1).date()
    #end_of_month = (pd.to_datetime('today').normalize() + pd.offsets.MonthEnd(0)).date()


    

    # Lista de portadores para a aba "Créditos e Devoluções"
    portadores_creditos_devolucoes = [
        'NOTA DE CREDITO', 'DEVOLUÇÃO DE VENDA', 'VALOR PAGO A MAIS', 'CREDITO POR TROCA',
        'CASHBACK', 'COMISSÃO PJ', 'VOUCHER SKY SETEMBRO', 'VOUCHER SKY OUTUBRO',
        'VOUCHER SKY NOVEMBRO', 'VOUCHER SKY MARÇO', 'VOUCHER SKY MAIO', 'VOUCHER SKY JUNHO',
        'VOUCHER SKY JULHO', 'VOUCHER SKY JANEIRO', 'VOUCHER SKY FEVEREIR', 'VOUCHER SKY DEZEMBRO',
        'VOUCHER SKY AGOSTO', 'VOUCHER SKY ABRIL'
    ]

    # Abas para tabelas com carregamento sob demanda
    st.subheader("Tabelas de Contas")
    show_legend = st.checkbox("Exibir Legenda", value=True)
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        f"Contas em Atraso até {today.strftime('%d/%m/%Y')}",
        f"Contas a Pagar Hoje {today.strftime('%d/%m/%Y')}",
        f"Contas da Semana {start_of_week.strftime('%d/%m/%Y')} - {end_of_week.strftime('%d/%m/%Y')}",
        f"Contas do Mês {start_of_month.strftime('%m/%Y')}",
        "Créditos e Devoluções"
    ])

    with tab1:
        overdue_data = filtered_data[
            (filtered_data['Status Pagamento'] == 'Não Pago - Em Atraso') &
            (~filtered_data['Portador'].isin(portadores_creditos_devolucoes))
        ]
        display_aggrid_table(f"Contas em Atraso até {today.strftime('%d/%m/%Y')}", overdue_data, key="aggrid_overdue")

    with tab2:
        today_data = filtered_data[
            (pd.to_datetime(filtered_data['Dt Vencimento']).dt.normalize() == pd.to_datetime(today).normalize()) &
            (~filtered_data['Portador'].isin(portadores_creditos_devolucoes))
        ]
        display_aggrid_table(f"Contas a Pagar Hoje {today.strftime('%d/%m/%Y')}", today_data, key="aggrid_today")

    with tab3:
        week_data = filtered_data[
            (pd.to_datetime(filtered_data['Dt Vencimento']).dt.date >= start_of_week) &
            (pd.to_datetime(filtered_data['Dt Vencimento']).dt.date <= end_of_week) &
            (~filtered_data['Portador'].isin(portadores_creditos_devolucoes))
        ]
        display_aggrid_table(f"Contas da Semana {start_of_week.strftime('%d/%m/%Y')} - {end_of_week.strftime('%d/%m/%Y')}", week_data, key="aggrid_week")

    with tab4:
        month_data = filtered_data[
            (pd.to_datetime(filtered_data['Dt Vencimento']).dt.date >= start_of_month) &
            (pd.to_datetime(filtered_data['Dt Vencimento']).dt.date <= end_of_month) &
            (~filtered_data['Portador'].isin(portadores_creditos_devolucoes))
        ]
        display_aggrid_table(f"Contas do Mês {start_of_month.strftime('%m/%Y')}", month_data, key="aggrid_month")

    with tab5:
        creditos_devolucoes_data = filtered_data[
            filtered_data['Portador'].isin(portadores_creditos_devolucoes)
        ]
        display_aggrid_table("Créditos e Devoluções", creditos_devolucoes_data, key="aggrid_creditos_devolucoes")

    st.write("Dados atualizados automaticamente a cada 30 minutos.")